import pandas as pd
import os
import datetime
import hashlib
import re
import queue
import threading
from .config import ConfigManager
from .id_registry import IDRegistry
from .utils import backup_excel_file
from .constants import (
    STATUS_IN, STATUS_OUT, STATUS_RETURN, STATUS_SOLD,
    ACTION_STATUS_CHANGE, ACTION_DATA_UPDATE, ACTION_MERGE, ACTION_REDIRECT,
    ACTION_ITEM_UPDATE,
    FIELD_IMEI, FIELD_MODEL, FIELD_PRICE, FIELD_STATUS, FIELD_BUYER, FIELD_BUYER_CONTACT,
    FIELD_UNIQUE_ID, FIELD_SOURCE_FILE, FIELD_NOTES, FIELD_COLOR, FIELD_RAM_ROM,
    FIELD_PRICE_ORIGINAL, ACTION_RELOAD
)

class InventoryManager:
    def __init__(self, config_manager: ConfigManager, activity_logger=None):
        self.config_manager = config_manager
        self.activity_logger = activity_logger
        self.id_registry = IDRegistry()
        self.inventory_df = pd.DataFrame()
        self.file_status = {}  # Keep track of file read status
        self.conflicts = []
        
        # Background Write Queue
        self.write_queue = queue.Queue()
        self._start_worker()

    def _start_worker(self):
        """Starts a background thread to handle Excel writes sequentially."""
        def worker():
            while True:
                task = self.write_queue.get()
                if task is None: break
                try:
                    func, args = task
                    func(*args)
                except Exception as e:
                    print(f"Background Worker Error: {e}")
                finally:
                    self.write_queue.task_done()
        
        t = threading.Thread(target=worker, daemon=True, name="ExcelWriterThread")
        t.start()

    def load_file(self, file_path):
        """Legacy wrapper or simple load."""
        mapping_data = self.config_manager.get_file_mapping(file_path)
        return self._load_file_internal(file_path, mapping_data)

    def _load_file_internal(self, file_path, mapping_data):
        if not mapping_data:
            return None, "MAPPING_REQUIRED"

        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
            else:
                # Support specific sheet name
                sheet_name = mapping_data.get('sheet_name', 0)
                # Handle None or empty string properly
                if sheet_name is None or sheet_name == "":
                    sheet_name = 0
                
                # Check if sheet exists before reading to avoid generic KeyError
                import openpyxl
                temp_wb = openpyxl.load_workbook(file_path, read_only=True)
                if isinstance(sheet_name, str) and sheet_name not in temp_wb.sheetnames:
                    temp_wb.close()
                    return None, f"SHEET_NOT_FOUND: {sheet_name}"
                temp_wb.close()

                df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Safety: Ensure DataFrame
            if isinstance(df, pd.Series):
                df = df.to_frame().T
            
            # Normalize columns based on mapping
            normalized_df = self._normalize_data(df, mapping_data, file_path)
            return normalized_df, "SUCCESS"
        except Exception as e:
            return None, str(e)

    def _normalize_data(self, df, mapping_data, file_path):
        """
        Converts the raw dataframe to the canonical format using the mapping.
        """
        mapping = mapping_data.get('mapping', {})
        file_supplier = mapping_data.get('supplier', '')
        
        # Helper: Normalize Status
        def norm_status(val):
            s = str(val).upper().strip()
            if s in [STATUS_OUT, 'SOLD', 'SALE']: return STATUS_OUT
            if s in [STATUS_RETURN, 'RETURN', 'RET']: return STATUS_RETURN
            if s in ['AVAILABLE', 'AVBL', 'STOCK', STATUS_IN, 'NAN', 'NONE', '', 'NAN']: return STATUS_IN
            return STATUS_IN

        # Create a new DF with canonical columns
        canonical = pd.DataFrame()
        
        # Helper to safely get column data as Series
        def get_col(canonical_name):
            mapped_col = None
            for col_name, can_name in mapping.items():
                if can_name == canonical_name:
                    mapped_col = col_name
                    break
            
            if mapped_col and mapped_col in df.columns:
                # SAFE CAST: Convert to object to avoid float.fillna errors
                return df[mapped_col].astype(object)
            return None

        # Map fields
        
        # IMEI Cleaning & Dual Support
        col_imei = get_col(FIELD_IMEI)
        if col_imei is not None:
            def clean_imei(val):
                if pd.isna(val): return ""
                s = str(val).strip()
                # Find sequences of 14-16 digits
                nums = re.findall(r'\d{14,16}', s)
                if len(nums) > 1:
                    return " / ".join(sorted(list(set(nums)))) # Unique, sorted
                elif len(nums) == 1:
                    return nums[0]
                return s
            canonical[FIELD_IMEI] = col_imei.apply(clean_imei)
        else:
            canonical[FIELD_IMEI] = ''
        
        # Model
        col_model = get_col(FIELD_MODEL)
        canonical[FIELD_MODEL] = col_model.fillna('Unknown Model').astype(str) if col_model is not None else 'Unknown Model'
        
        # Brand (Extract from model if not mapped)
        col_brand = get_col('brand')
        if col_brand is not None:
            canonical['brand'] = col_brand.fillna('').astype(str).str.upper()
        else:
            # Fallback: First word of model
            canonical['brand'] = canonical[FIELD_MODEL].apply(lambda x: str(x).split()[0].upper() if x and str(x).split() else 'UNKNOWN')

        # Price Logic
        col_price = get_col(FIELD_PRICE)
        if col_price is not None:
            raw_price = pd.to_numeric(col_price, errors='coerce').fillna(0.0)
        else:
            raw_price = 0.0
        
        canonical[FIELD_PRICE_ORIGINAL] = raw_price
        
        # Apply Markup
        try:
            markup = float(self.config_manager.get('price_markup_percent', 0.0))
        except (ValueError, TypeError):
            markup = 0.0
            
        if markup > 0:
            # Vectorized calc
            price_with_markup = raw_price * (1 + markup/100.0)
            # Round to nearest 100: round(x/100)*100
            canonical[FIELD_PRICE] = price_with_markup.apply(lambda x: round(x / 100) * 100 if x > 0 else x)
        else:
            canonical[FIELD_PRICE] = raw_price
        
        # RAM/ROM handling
        ram_rom = get_col(FIELD_RAM_ROM)
        if ram_rom is not None:
             canonical[FIELD_RAM_ROM] = ram_rom.fillna('').astype(str)
        else:
            ram = get_col('ram')
            rom = get_col('rom')
            if ram is not None and rom is not None:
                canonical[FIELD_RAM_ROM] = ram.fillna('').astype(str) + " / " + rom.fillna('').astype(str)
            elif ram is not None:
                canonical[FIELD_RAM_ROM] = ram.fillna('').astype(str)
            else:
                canonical[FIELD_RAM_ROM] = ""

        # Supplier handling
        supplier_col = get_col('supplier')
        if supplier_col is not None:
            canonical['supplier'] = supplier_col.fillna(file_supplier).astype(str)
        else:
            canonical['supplier'] = file_supplier

        # Status handling (Excel)
        status_col = get_col(FIELD_STATUS)
        if status_col is not None:
            canonical[FIELD_STATUS] = status_col.apply(norm_status)
        else:
            canonical[FIELD_STATUS] = STATUS_IN # Default status (Available)

        # Color handling
        color_col = get_col(FIELD_COLOR)
        if color_col is not None:
            canonical[FIELD_COLOR] = color_col.fillna('').astype(str)
        else:
            canonical[FIELD_COLOR] = ''

        # Buyer handling
        col_buyer = get_col(FIELD_BUYER)
        canonical[FIELD_BUYER] = col_buyer.fillna('').astype(str) if col_buyer is not None else ''
        
        col_contact = get_col(FIELD_BUYER_CONTACT)
        canonical[FIELD_BUYER_CONTACT] = col_contact.fillna('').astype(str) if col_contact is not None else ''

        # Grade/Condition
        col_grade = get_col('grade')
        canonical['grade'] = col_grade.fillna('').astype(str) if col_grade is not None else ''

        col_condition = get_col('condition')
        canonical['condition'] = col_condition.fillna('').astype(str) if col_condition is not None else ''

        # Metadata
        canonical[FIELD_SOURCE_FILE] = str(file_path)
        canonical['last_updated'] = datetime.datetime.now()
        
        # Ensure all required columns exist with defaults
        required_cols = [
            FIELD_UNIQUE_ID, FIELD_IMEI, 'brand', FIELD_MODEL, FIELD_RAM_ROM,
            FIELD_PRICE, FIELD_PRICE_ORIGINAL, 'supplier', FIELD_SOURCE_FILE,
            'last_updated', FIELD_STATUS, FIELD_COLOR, FIELD_NOTES, FIELD_BUYER,
            FIELD_BUYER_CONTACT, 'grade', 'condition'
        ]
        for col in required_cols:
            if col not in canonical.columns:
                canonical[col] = "" if col != FIELD_PRICE else 0.0

        # --- VECTORIZED ID GENERATION ---
        # Generate keys for all rows at once
        has_imei = (canonical[FIELD_IMEI].str.len() > 4)
        keys = pd.Series("", index=canonical.index)
        
        # IMEI Keys
        keys[has_imei] = "IMEI:" + canonical.loc[has_imei, FIELD_IMEI].str.strip()
        
        # Hash Keys for items without valid IMEI
        no_imei = ~has_imei
        if no_imei.any():
            combined = (
                canonical.loc[no_imei, FIELD_MODEL].astype(str) + "|" +
                canonical.loc[no_imei, FIELD_RAM_ROM].astype(str) + "|" +
                canonical.loc[no_imei, 'supplier'].astype(str)
            )
            keys[no_imei] = "HASH:" + combined.apply(lambda x: hashlib.md5(x.encode()).hexdigest())
        
        # Batch lookup/create IDs
        canonical[FIELD_UNIQUE_ID] = self.id_registry.get_ids_batch(keys.tolist())
        
        # Merge persistent metadata (status, notes, color, price overrides)
        # Pre-fetch metadata to avoid repeated dict lookups in apply
        uids = canonical[FIELD_UNIQUE_ID].unique()
        metadata_map = {str(uid): self.id_registry.get_metadata(uid) for uid in uids}
        
        def apply_overrides(row):
            uid_str = str(row[FIELD_UNIQUE_ID])
            meta = metadata_map.get(uid_str, {})
            if not meta: return row
            
            # Use app-stored status if present
            if FIELD_STATUS in meta:
                row[FIELD_STATUS] = norm_status(meta[FIELD_STATUS])
                
            if FIELD_NOTES in meta: row[FIELD_NOTES] = meta[FIELD_NOTES]
            if FIELD_COLOR in meta: row[FIELD_COLOR] = meta[FIELD_COLOR]
            if 'grade' in meta: row['grade'] = meta['grade']
            if 'condition' in meta: row['condition'] = meta['condition']
            
            if FIELD_PRICE_ORIGINAL in meta:
                row[FIELD_PRICE_ORIGINAL] = float(meta[FIELD_PRICE_ORIGINAL])
                try:
                    markup = float(self.config_manager.get('price_markup_percent', 0.0))
                except:
                    markup = 0.0
                    
                if markup > 0:
                    raw_p = row[FIELD_PRICE_ORIGINAL] * (1 + markup/100.0)
                    row[FIELD_PRICE] = round(raw_p / 100) * 100 if raw_p > 0 else raw_p
                else:
                    row[FIELD_PRICE] = row[FIELD_PRICE_ORIGINAL]
            return row
            
        canonical = canonical.apply(apply_overrides, axis=1)
        return canonical

    def reload_all(self):
        """Reloads all files in mappings and merges them."""
        all_frames = []
        mappings = self.config_manager.mappings
        self.conflicts = [] # Reset conflicts
        
        # Performance optimization: Disable auto-save during batch load
        self.id_registry.auto_save = False
        
        try:
            for key, mapping_data in mappings.items():
                # Support composite keys "path::sheet" or legacy "path"
                file_path = mapping_data.get('file_path', key)
                
                # Fallback if file_path is not in data and key looks composite
                if '::' in key and not os.path.exists(key):
                    parts = key.split('::')
                    if os.path.exists(parts[0]):
                        file_path = parts[0]
                
                if os.path.exists(file_path):
                    df, status = self._load_file_internal(file_path, mapping_data)

                    if status == "SUCCESS" and df is not None:
                        df[FIELD_SOURCE_FILE] = key 
                        all_frames.append(df)
                        self.file_status[key] = "OK"
                    else:
                        self.file_status[key] = f"Error: {status}"
                else:
                    self.file_status[key] = "Missing"
            
            if all_frames:
                full_df = pd.concat(all_frames, ignore_index=True)
                
                # Filter Hidden Items (Merge Logic)
                # Optimization: Vectorized filter
                uids = full_df[FIELD_UNIQUE_ID].astype(str)
                hidden_ids = {iid for iid, meta in self.id_registry.registry.get('metadata', {}).items() if meta.get('is_hidden')}
                full_df = full_df[~uids.isin(hidden_ids)]
                
                # Detect Duplicates (Optimized)
                # Only check items that have IMEIs
                imei_df = full_df[full_df[FIELD_IMEI].str.len() > 5].copy()
                if not imei_df.empty:
                    # Explode dual IMEIs for easier detection
                    # "IMEI1 / IMEI2" -> ["IMEI1", "IMEI2"]
                    imei_df['imei_list'] = imei_df[FIELD_IMEI].str.split('/')
                    exploded = imei_df.explode('imei_list')
                    exploded['imei_list'] = exploded['imei_list'].str.strip()
                    exploded = exploded[exploded['imei_list'].str.len() > 5]
                    
                    dupes = exploded[exploded.duplicated('imei_list', keep=False)]
                    for imei, group in dupes.groupby('imei_list'):
                        self.conflicts.append({
                            "imei": imei,
                            "unique_ids": group[FIELD_UNIQUE_ID].unique().tolist(),
                            "model": group.iloc[0][FIELD_MODEL],
                            "sources": group[FIELD_SOURCE_FILE].unique().tolist(),
                            "rows": group.drop_duplicates(FIELD_UNIQUE_ID).to_dict('records')
                        })
                
                self.inventory_df = full_df
            else:
                self.inventory_df = pd.DataFrame(columns=[
                    FIELD_UNIQUE_ID, FIELD_IMEI, 'brand', FIELD_MODEL, FIELD_RAM_ROM, 
                    FIELD_PRICE, FIELD_PRICE_ORIGINAL, 'supplier', FIELD_SOURCE_FILE, 
                    'last_updated', FIELD_STATUS, FIELD_COLOR, FIELD_NOTES, FIELD_BUYER,
                    FIELD_BUYER_CONTACT, 'grade', 'condition'
                ])
        finally:
            self.id_registry.commit()
            self.id_registry.auto_save = True
            
        if self.activity_logger:
            self.activity_logger.log(ACTION_RELOAD, f"Loaded {len(self.inventory_df)} items from {len(mappings)} sources.")
            
        return self.inventory_df
    
    def resolve_conflict(self, conflict_data, action, keep_source=None):
        """
        Resolves an IMEI conflict.
        action: 'merge' (hides duplicates)
        """
        if action == 'merge':
            rows = conflict_data.get('rows', [])
            if not rows: return False
            
            # Strategy: Keep the first one (or specific if logic allows), hide others
            keeper = rows[0]
            others = rows[1:]
            
            for row in others:
                uid = row[FIELD_UNIQUE_ID]
                
                # CRITICAL FIX: Do not merge if IDs are identical (Self-Merge)
                # This prevents an item from hiding itself.
                if str(uid) == str(keeper[FIELD_UNIQUE_ID]):
                    continue
                    
                self.id_registry.update_metadata(uid, {
                    'is_hidden': True, 
                    'merged_into': keeper[FIELD_UNIQUE_ID],
                    'merge_reason': 'Conflict Resolution'
                })
                # Log
                self.id_registry.add_history_log(uid, ACTION_MERGE, f"Merged into {keeper[FIELD_UNIQUE_ID]}")
                if self.activity_logger:
                    self.activity_logger.log(ACTION_MERGE, f"Merged {uid} into {keeper[FIELD_UNIQUE_ID]}")
            
            self.reload_all()
            return True
            
        return False

    def get_merged_target(self, unique_id):
        """
        Checks if an ID is merged/hidden and returns the Target ID it was merged into.
        Returns None if not merged.
        """
        meta = self.id_registry.get_metadata(unique_id)
        if meta.get('is_hidden', False):
            return meta.get('merged_into')
        return None

    def get_item_by_id(self, unique_id, resolve_merged=True):
        """
        Finds an item by ID. 
        If resolve_merged=True, it follows 'merged_into' links.
        Returns (ItemDict, RedirectedID) or (None, None).
        """
        unique_id = str(unique_id).strip()
        
        # 1. Check if merged
        target_id = unique_id
        if resolve_merged:
            merged_into = self.get_merged_target(unique_id)
            if merged_into:
                target_id = str(merged_into)
        
        # 2. Lookup in Inventory DF
        df = self.inventory_df
        mask = df[FIELD_UNIQUE_ID].astype(str) == target_id
        
        if mask.any():
            item = df[mask].iloc[0].to_dict()
            redirected_from = unique_id if target_id != unique_id else None
            return item, redirected_from
            
        return None, None

    def get_inventory(self):
        return self.inventory_df

    def update_item_status(self, item_id, new_status, write_to_excel=False):
        """Updates and persists the status of an item."""
        # --- REDIRECT MERGED IDs ---
        target_id = self.get_merged_target(item_id)
        if target_id:
            if self.activity_logger:
                self.activity_logger.log(ACTION_REDIRECT, f"Status update for {item_id} redirected to {target_id}")
            item_id = target_id

        # 1. Get Previous Status for logging
        mask = self.inventory_df[FIELD_UNIQUE_ID].astype(str) == str(item_id)
        old_status = "UNKNOWN"
        item_model = ""
        if mask.any():
            row = self.inventory_df[mask].iloc[0]
            old_status = row[FIELD_STATUS]
            item_model = row[FIELD_MODEL]

        # 2. Update Internal Registry
        self.id_registry.update_metadata(item_id, {FIELD_STATUS: new_status})
        self.id_registry.add_history_log(item_id, ACTION_STATUS_CHANGE, f"Moved from {old_status} to {new_status}")
        
        if self.activity_logger:
            self.activity_logger.log(ACTION_STATUS_CHANGE, f"Item {item_id} ({item_model}) marked as {new_status}")
        
        # 3. Update Memory
        if mask.any():
            self.inventory_df.loc[mask, FIELD_STATUS] = new_status
            
            # 3. Write to Excel (ASYNC via Queue)
            if write_to_excel:
                try:
                    # Create a snapshot of the row data
                    row_snapshot = self.inventory_df[mask].iloc[0].to_dict()
                    updates = {FIELD_STATUS: new_status}
                    
                    # Offload to worker thread
                    self.write_queue.put((self._write_excel_generic, (row_snapshot, updates)))
                    return True
                except Exception as e:
                    print(f"Queue Error: {e}")
                    return False
        return True

    def update_item_data(self, item_id, updates):
        """Updates generic item data (price, color, etc) and writes to Excel."""
        # --- REDIRECT MERGED IDs ---
        target_id = self.get_merged_target(item_id)
        if target_id:
            item_id = target_id

        mask = self.inventory_df[FIELD_UNIQUE_ID].astype(str) == str(item_id)
        if not mask.any(): return False
        
        item_model = self.inventory_df.loc[mask, FIELD_MODEL].values[0]
        
        # 1. Update Persistent Registry (App DB)
        self.id_registry.update_metadata(item_id, updates)
        
        # Log generic updates
        log_details = ", ".join([f"{k}={v}" for k, v in updates.items()])
        self.id_registry.add_history_log(item_id, ACTION_DATA_UPDATE, log_details)
        
        if self.activity_logger:
            self.activity_logger.log(ACTION_ITEM_UPDATE, f"Updated {item_id} ({item_model}): {log_details}")
        
        # 2. Update Memory
        for k, v in updates.items():
            if k in self.inventory_df.columns:
                self.inventory_df.loc[mask, k] = v
        
        # 3. Write to Excel (ASYNC via Queue)
        try:
            row_snapshot = self.inventory_df[mask].iloc[0].to_dict()
            self.write_queue.put((self._write_excel_generic, (row_snapshot, updates)))
            return True
        except Exception as e:
            print(f"Queue error: {e}")
            return False

    def _write_excel_generic(self, row_data, updates):
        # NOTE: This runs in a background thread!
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from copy import copy
        
        # Handle composite keys (path::sheet)
        key = row_data[FIELD_SOURCE_FILE]
        file_path = key
        if '::' in key and not os.path.exists(key):
            file_path = key.split('::')[0]

        if not os.path.exists(file_path): 
            print(f"Write Error: File not found {file_path}")
            return False, f"Source file not found: {file_path}"

        # SAFETY 1: Backup
        backup_path = backup_excel_file(file_path)
        if not backup_path:
            print("Write Error: Backup failed")
            return False, "Failed to create backup, aborting write for safety."
            
        # Use the original KEY to get mapping data (it might be keyed by "path::sheet")
        mapping_data = self.config_manager.get_file_mapping(key)
        # If fallback needed (mapping might be keyed by just path)
        if not mapping_data:
             mapping_data = self.config_manager.get_file_mapping(file_path)
             
        mapping = mapping_data.get('mapping', {})
        
        # Build map of InternalField -> ExcelColumnName
        field_to_col = {v: k for k, v in mapping.items()}
        
        # Map updates to Excel headers
        excel_updates = {}
        for k, v in updates.items():
            if k == FIELD_PRICE_ORIGINAL: k = FIELD_PRICE
            if k in field_to_col:
                excel_updates[field_to_col[k]] = v
        
        max_retries = 3
        retry_delay = 1.5 # seconds
        
        for attempt in range(max_retries):
            try:
                wb = load_workbook(file_path)
                
                # Determine Sheet
                sheet_name = mapping_data.get('sheet_name', 0)
                if '::' in key:
                     try:
                         sheet_part = key.split('::')[1]
                         if sheet_part: 
                             # Try converting to int if it looks like one
                             if sheet_part.isdigit():
                                 sheet_name = int(sheet_part)
                             else:
                                 sheet_name = sheet_part
                     except: pass

                ws = None
                # 1. Try Integer Index
                if isinstance(sheet_name, int):
                    try:
                        ws = wb.worksheets[sheet_name]
                    except IndexError:
                        print(f"Sheet index {sheet_name} out of range, falling back to active.")
                
                # 2. Try String Name
                elif isinstance(sheet_name, str):
                     if sheet_name in wb.sheetnames:
                         ws = wb[sheet_name]
                     elif sheet_name.isdigit(): # "0" as string?
                         try:
                             ws = wb.worksheets[int(sheet_name)]
                         except: pass

                # 3. Fallback
                if ws is None:
                    ws = wb.active
                    
                print(f"BG-WRITE: Writing to Sheet '{ws.title}' in '{file_path}' (Attempt {attempt+1})")

                # Find Header Columns
                col_indices = {}
                max_col_idx = 0
                for cell in ws[1]:
                    val = str(cell.value).strip()
                    if val:
                        col_indices[val] = cell.column
                        if cell.column > max_col_idx: max_col_idx = cell.column
                
                default_headers = {
                    FIELD_BUYER: 'Buyer Name',
                    FIELD_BUYER_CONTACT: 'Buyer Contact',
                    FIELD_NOTES: 'Notes',
                    FIELD_STATUS: 'Status',
                    FIELD_COLOR: 'Color',
                    FIELD_PRICE: 'Selling Price',
                    'grade': 'Grade',
                    'condition': 'Condition'
                }
                
                for field, value in updates.items():
                    header_name = None
                    if field in field_to_col:
                        header_name = field_to_col[field]
                    elif field in default_headers:
                        header_name = default_headers[field]
                    
                    if header_name:
                        if header_name not in col_indices:
                            # Create New Column
                            max_col_idx += 1
                            ws.cell(row=1, column=max_col_idx).value = header_name
                            col_indices[header_name] = max_col_idx
                        excel_updates[header_name] = value

                # Find Row (Match IMEI or Model)
                target_imei = str(row_data[FIELD_IMEI])
                target_model = str(row_data[FIELD_MODEL])
                
                # Find IMEI/Model col for matching
                imei_header = field_to_col.get(FIELD_IMEI)
                model_header = field_to_col.get(FIELD_MODEL)
                
                imei_col_idx = col_indices.get(imei_header)
                model_col_idx = col_indices.get(model_header)
                
                row_found = False
                
                for row in ws.iter_rows(min_row=2):
                    match = False
                    if imei_col_idx:
                        cell_val = row[imei_col_idx-1].value
                        if cell_val:
                            s_cell = str(cell_val).strip().replace('.0', '')
                            # Robust Check: Exact match or partial (for dual IMEI scenarios)
                            # target_imei might be "A / B", cell might be "A" or "B" or "A/B"
                            if s_cell == target_imei:
                                match = True
                            elif target_imei and (target_imei in s_cell or s_cell in target_imei):
                                match = True
                    
                    if not match and model_col_idx:
                        if str(row[model_col_idx-1].value).strip() == target_model:
                            match = True # Weak match fallback
                    
                    if match:
                        row_found = True
                        # Apply updates
                        for col_name, new_val in excel_updates.items():
                            if col_name in col_indices:
                                cell = ws.cell(row=row[0].row, column=col_indices[col_name])
                                
                                # Enforce Uppercase for strings
                                if isinstance(new_val, str):
                                    new_val = new_val.upper()
                                    
                                cell.value = new_val
                                
                                # --- ENFORCE USER STYLE ---
                                # Times New Roman, 11, Bold, Center, All Borders
                                thin = Side(border_style="thin", color="000000")
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                                cell.font = Font(name='Times New Roman', size=11, bold=True)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                    
                        break
                
                if not row_found:
                    print(f"Warning: No matching row found in Excel for {target_imei} / {target_model}")
                    return False, f"Row not found for {target_imei}"
                    
                wb.save(file_path)
                return True, "Success"
                
            except PermissionError:
                print(f"Write Attempt {attempt+1} failed: File open in Excel. Retrying...")
                if attempt < max_retries - 1:
                    import time
                    time.sleep(retry_delay)
                else:
                    return False, "File is open in Excel. Please close it."
            except Exception as e:
                print(f"Write Error: {e}")
                return False, f"Excel Write Error: {e}"